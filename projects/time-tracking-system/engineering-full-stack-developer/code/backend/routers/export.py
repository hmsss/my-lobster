"""导出相关API路由"""
import io
import csv
from fastapi import APIRouter, Query, Depends
from fastapi.responses import StreamingResponse
from database import get_db
from schemas import ApiResponse
from auth import get_current_admin_user
from config import MAX_EXPORT_ROWS

router = APIRouter()


@router.get("")
async def export_timesheet(
    format: str = Query("excel"),
    startDate: str = Query(None),
    endDate: str = Query(None),
    userIds: str = Query(None),
    projectIds: str = Query(None),
    current_user: dict = Depends(get_current_admin_user)
):
    """导出工时数据（仅管理员）"""
    with get_db() as conn:
        # 构建查询条件
        conditions = []
        params = []
        
        if startDate:
            conditions.append("te.work_date >= ?")
            params.append(startDate)
        if endDate:
            conditions.append("te.work_date <= ?")
            params.append(endDate)
        if userIds:
            user_id_list = [int(uid) for uid in userIds.split(",")]
            placeholders = ",".join("?" * len(user_id_list))
            conditions.append(f"te.user_id IN ({placeholders})")
            params.extend(user_id_list)
        if projectIds:
            project_id_list = [int(pid) for pid in projectIds.split(",")]
            placeholders = ",".join("?" * len(project_id_list))
            conditions.append(f"te.project_id IN ({placeholders})")
            params.extend(project_id_list)
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        
        # 查询数据
        rows = conn.execute(f"""
            SELECT u.name as user_name, te.work_date, p.name as project_name,
                   te.task, te.hours, te.note, te.created_at
            FROM time_entries te
            JOIN users u ON te.user_id = u.id
            JOIN projects p ON te.project_id = p.id
            WHERE {where_clause}
            ORDER BY te.work_date DESC, u.name
            LIMIT ?
        """, params + [MAX_EXPORT_ROWS]).fetchall()
        
        if not rows:
            return ApiResponse(code=4001, message="导出数据为空")
        
        if format == "csv":
            # CSV 导出
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["员工姓名", "日期", "项目", "任务", "时长(小时)", "备注", "提交时间"])
            
            for row in rows:
                writer.writerow([
                    row["user_name"],
                    row["work_date"],
                    row["project_name"],
                    row["task"],
                    row["hours"],
                    row["note"] or "",
                    row["created_at"]
                ])
            
            output.seek(0)
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode("utf-8-sig")),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=timesheet.csv"}
            )
        else:
            # Excel 导出
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "工时数据"
            
            # 表头
            headers = ["员工姓名", "日期", "项目", "任务", "时长(小时)", "备注", "提交时间"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # 数据
            for row_idx, row in enumerate(rows, 2):
                ws.cell(row=row_idx, column=1, value=row["user_name"])
                ws.cell(row=row_idx, column=2, value=row["work_date"])
                ws.cell(row=row_idx, column=3, value=row["project_name"])
                ws.cell(row=row_idx, column=4, value=row["task"])
                ws.cell(row=row_idx, column=5, value=row["hours"])
                ws.cell(row=row_idx, column=6, value=row["note"] or "")
                ws.cell(row=row_idx, column=7, value=row["created_at"])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 20
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=timesheet.xlsx"}
            )
